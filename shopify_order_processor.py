# -*- coding: utf-8 -*-
"""
Shopify Order Processor

This script processes Shopify order exports (CSV) to filter, aggregate,
and format them into a clean Excel report.
"""
import sys
import os
import logging
from datetime import datetime

import pandas as pd
# openpyxl is used by pandas for Excel writing, so we import it to ensure it's available.
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter

def validate_date_format(date_string):
    """
    Validates that a date string is in the DD.MM.YYYY format.

    Args:
        date_string (str): The string to validate.

    Returns:
        datetime.datetime | None: A datetime object if the format is correct,
                                  otherwise None.
    """
    try:
        return datetime.strptime(date_string, "%d.%m.%Y")
    except ValueError:
        return None

def get_tariff_from_user(prompt_message):
    """
    Prompts the user for a numeric tariff value and validates it.

    The function will loop indefinitely until the user enters a non-negative
    number.

    Args:
        prompt_message (str): The message to display to the user.

    Returns:
        float: The validated, non-negative tariff value.
    """
    while True:
        cost_str = input(prompt_message)
        try:
            cost = float(cost_str)
            if cost >= 0:
                return cost
            else:
                logging.error("Please enter a non-negative number.")
        except ValueError:
            logging.error("Invalid input. Please enter a valid number (e.g., 10.50).")

def get_date_from_user(prompt_message):
    """
    Prompts the user for a date and validates it is in DD.MM.YYYY format.

    The function will loop indefinitely until the user enters a valid date
    string.

    Args:
        prompt_message (str): The message to display to the user.

    Returns:
        datetime.datetime: The validated date as a datetime object.
    """
    while True:
        date_str = input(prompt_message)
        date_obj = validate_date_format(date_str)
        if date_obj:
            return date_obj
        else:
            logging.error("Invalid date format. Please use DD.MM.YYYY.")

def load_and_validate_csv(file_path):
    """
    Loads a CSV file, validates its existence and required columns, and
    cleans the 'Name' column to ensure consistent grouping.
    """
    if not os.path.exists(file_path):
        logging.error(f"The file '{file_path}' was not found.")
        sys.exit(1)

    dtype_map = {
        'Subtotal': float, 'Shipping': float, 'Taxes': float, 'Total': float,
        'Discount Amount': float, 'Lineitem quantity': int, 'Lineitem price': float,
        'Lineitem compare at price': float, 'Lineitem requires shipping': bool,
        'Lineitem taxable': bool, 'Refunded Amount': float, 'Outstanding Balance': float,
        'Id': float, 'Lineitem discount': float, 'Tax 1 Value': float, 'Tax 2 Value': float,
        'Tax 3 Value': float, 'Tax 4 Value': float, 'Tax 5 Value': float, 'Duties': float,
    }

    try:
        df = pd.read_csv(file_path, dtype=dtype_map, low_memory=False)
    except Exception as e:
        logging.error(f"Failed to read the CSV file. Reason: {e}")
        sys.exit(1)

    required_columns = [
        'Name', 'Fulfilled at', 'Lineitem quantity',
        'Lineitem name', 'Lineitem sku', 'Total'
    ]
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        logging.error(f"The input CSV is missing the following required columns: {', '.join(missing_columns)}")
        sys.exit(1)

    # --- Data Cleaning Step ---
    # Strip whitespace from 'Name' column to prevent grouping errors
    if 'Name' in df.columns:
        df['Name'] = df['Name'].astype(str).str.strip()


    logging.info("Successfully loaded and validated the input file.")
    return df

def filter_by_date_range(df, start_date, end_date):
    """
    Filters a DataFrame of orders by a given date range.

    The filtering is based on the 'Fulfilled at' column. The function handles
    missing fulfillment dates, date parsing errors, and timezone conversion.

    Args:
        df (pd.DataFrame): The input DataFrame of orders.
        start_date (datetime.datetime): The start of the filtering period.
        end_date (datetime.datetime): The end of the filtering period.

    Returns:
        pd.DataFrame: A new DataFrame containing only the orders within the
                      specified date range.
    """
    logging.info(f"Initial record count: {len(df)}")

    ffill_cols = ['Fulfilled at', 'Fulfillment Status', 'Financial Status']
    for col in ffill_cols:
        if col in df.columns:
            df[col] = df.groupby('Name')[col].ffill()

    original_count = len(df)
    df.dropna(subset=['Fulfilled at'], inplace=True)
    if len(df) < original_count:
        logging.info(f"Dropped {original_count - len(df)} unfulfilled orders.")

    df['Fulfilled at'] = pd.to_datetime(df['Fulfilled at'], errors='coerce')

    original_count = len(df)
    df.dropna(subset=['Fulfilled at'], inplace=True)
    if len(df) < original_count:
        logging.warning(f"Dropped {original_count - len(df)} rows with invalid date format in 'Fulfilled at'.")

    if pd.api.types.is_datetime64_any_dtype(df['Fulfilled at']) and df['Fulfilled at'].dt.tz is not None:
        df['Fulfilled at'] = df['Fulfilled at'].dt.tz_localize(None)

    mask = (df['Fulfilled at'].dt.normalize() >= start_date) & (df['Fulfilled at'].dt.normalize() <= end_date)
    filtered_df = df.loc[mask].copy()

    logging.info(f"Record count after filtering by date: {len(filtered_df)}")
    if len(filtered_df) == 0:
        logging.warning("No orders found within the specified date range.")
    return filtered_df

def calculate_costs(df, cost_first_sku, cost_next_sku, cost_per_piece):
    """
    Calculates processing costs for each line item in each order.

    This function iterates through each order, calculates costs on a line-by-line
    basis, and adds new columns for detailed cost breakdown.

    Args:
        df (pd.DataFrame): The DataFrame of orders.
        cost_first_sku (float): The cost for the first unique SKU in an order.
        cost_next_sku (float): The cost for each subsequent unique SKU.
        cost_per_piece (float): The cost for each individual item.

    Returns:
        pd.DataFrame: The DataFrame with added columns for line-item costs.
    """
    if df.empty:
        return df.copy()

    # Initialize new cost columns
    df['Piece Cost'] = 0.0
    df['SKU Cost'] = 0.0
    df['Line Total Cost'] = 0.0

    # Pattern to exclude non-billable items like insurance/protection
    protection_pattern = "Package protection|Shipping Protection"

    # Use a copy to avoid SettingWithCopyWarning when modifying groups
    df_copy = df.copy()

    # Iterate over each order group
    for name, group in df_copy.groupby('Name'):
        seen_skus = set()
        is_first_billable_item = True

        # Filter out non-billable items for cost calculation
        billable_items = group[~group['Lineitem name'].str.contains(protection_pattern, na=False)]

        for index, row in billable_items.iterrows():
            # 1. Calculate Piece Cost
            piece_cost = row['Lineitem quantity'] * cost_per_piece
            df.loc[index, 'Piece Cost'] = piece_cost

            # 2. Calculate SKU Cost
            sku_cost = 0.0
            if is_first_billable_item:
                sku_cost = cost_first_sku
                is_first_billable_item = False
            elif row['Lineitem sku'] not in seen_skus:
                sku_cost = cost_next_sku

            df.loc[index, 'SKU Cost'] = sku_cost
            seen_skus.add(row['Lineitem sku'])

            # 3. Calculate Total Line Cost
            df.loc[index, 'Line Total Cost'] = piece_cost + sku_cost

    return df

def create_invoice_summary(df_with_costs, cost_first_sku, cost_next_sku, cost_per_piece):
    """
    Creates a summary DataFrame formatted as a final invoice, based on line-item costs.
    """
    if df_with_costs.empty:
        return pd.DataFrame()

    protection_pattern = "Package protection|Shipping Protection"
    billable_df = df_with_costs[~df_with_costs['Lineitem name'].str.contains(protection_pattern, na=False)].copy()

    if billable_df.empty:
        return pd.DataFrame()

    total_orders = billable_df['Name'].nunique()
    if total_orders == 0:
        return pd.DataFrame()

    total_pieces = billable_df['Lineitem quantity'].sum()

    # Calculate SKU counts based on the 'SKU Cost' column values
    total_first_skus = (billable_df['SKU Cost'] == cost_first_sku).sum()
    total_next_skus = (billable_df['SKU Cost'] == cost_next_sku).sum()

    # Calculate total costs from the new columns
    total_sku_cost_calculated = billable_df['SKU Cost'].sum()
    total_piece_cost_calculated = billable_df['Piece Cost'].sum()
    grand_total_cost = billable_df['Line Total Cost'].sum()

    summary_data = {
        'Description': [
            'Total Orders Processed', 'First SKU Tariff', 'Subsequent SKU Tariff', 'Per-Piece Tariff',
            '---', 'Total SKU Cost', 'Total Piece Cost', '---', 'GRAND TOTAL'
        ],
        'Rate': [
            '', f'{cost_first_sku:.2f}', f'{cost_next_sku:.2f}', f'{cost_per_piece:.2f}',
            '', '', '', '', ''
        ],
        'Count': [
            total_orders, total_first_skus, total_next_skus, total_pieces,
            '', '', '', '', ''
        ],
        'Total Amount': [
            '', total_first_skus * cost_first_sku, total_next_skus * cost_next_sku, total_pieces * cost_per_piece,
            '', f'{total_sku_cost_calculated:.2f}', f'{total_piece_cost_calculated:.2f}', '', f'{grand_total_cost:.2f}'
        ]
    }
    return pd.DataFrame(summary_data)

def transform_cost_df_for_reporting(df_costs):
    """
    Transforms the cost calculation DataFrame to include a summary 'TOTAL' row for each order.
    This function uses a robust, vectorized approach to avoid ordering issues.
    """
    if df_costs.empty:
        return df_costs

    # 1. Calculate all totals first using a vectorized groupby.
    totals_df = df_costs.groupby('Name', as_index=False).agg(
        **{
            'Piece Cost': pd.NamedAgg(column='Piece Cost', aggfunc='sum'),
            'SKU Cost': pd.NamedAgg(column='SKU Cost', aggfunc='sum'),
            'Line Total Cost': pd.NamedAgg(column='Line Total Cost', aggfunc='sum'),
        }
    )
    totals_df['Lineitem name'] = 'TOTAL'

    # 2. Combine the original data with the new totals.
    # `sort=False` is not strictly needed here but is good practice.
    combined_df = pd.concat([df_costs, totals_df], ignore_index=True, sort=False)

    # 3. Create a temporary sort key to guarantee TOTAL rows appear last.
    # The 'is_total' column will be 0 for items and 1 for TOTALs.
    combined_df['is_total'] = (combined_df['Lineitem name'] == 'TOTAL').astype(int)

    # 4. Sort by the order name first, then by the 'is_total' flag.
    # This robustly places the TOTAL row at the bottom of each group.
    # We also fill NaNs in 'Fulfilled at' to prevent sorting errors if a TOTAL
    # row (with NaT) is compared with a valid date.
    if 'Fulfilled at' in combined_df.columns:
        # Use a far-future date for TOTAL rows to ensure they sort last if needed,
        # though sorting by 'is_total' should be sufficient.
        combined_df['Fulfilled at'].fillna(pd.Timestamp.max, inplace=True)

    final_df = combined_df.sort_values(by=['Name', 'is_total']).reset_index(drop=True)

    # 5. Clean up the temporary sort column.
    final_df = final_df.drop(columns=['is_total'])

    # Restore NaNs for presentation
    if 'Fulfilled at' in final_df.columns:
        final_df['Fulfilled at'] = final_df['Fulfilled at'].replace({pd.Timestamp.max: pd.NaT})


    return final_df

def create_excel_report(sheets_data, output_filename):
    """
    Creates a multi-sheet Excel report from a dictionary of DataFrames.
    """
    if not sheets_data:
        logging.info("No data to save, skipping Excel report generation.")
        return

    thin_side = Side(border_style="thin", color="000000")
    thick_side = Side(border_style="thick", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                if df.empty: continue
                # IMPORTANT: Do NOT re-sort 'Cost Calculation' here.
                # It is pre-sorted to ensure TOTAL rows are last.
                if sheet_name in ['All Orders', 'Without Package Protection']:
                    df = df.sort_values(by='Name').reset_index(drop=True)
                df.to_excel(writer, index=False, sheet_name=sheet_name)

            for sheet_name, worksheet in writer.sheets.items():
                if worksheet.max_row <= 1: continue

                header_font = Font(bold=True)
                for cell in worksheet["1:1"]:
                    cell.font = header_font
                    cell.border = thin_border
                worksheet.freeze_panes = 'A2'
                worksheet.auto_filter.ref = worksheet.dimensions

                df_to_format = sheets_data[sheet_name]
                # Re-sorting is not needed here either as the data is already in its final order.
                # if sheet_name in ['All Orders', 'Without Package Protection', 'Cost Calculation']:
                #      df_to_format = df_to_format.sort_values(by='Name').reset_index(drop=True)

                for i, col_name in enumerate(df_to_format.columns, 1):
                    column_letter = get_column_letter(i)
                    if col_name == 'Fulfilled at':
                        worksheet.column_dimensions[column_letter].width = 20
                        for cell in worksheet[column_letter][1:]:
                            if cell.value: cell.number_format = 'DD.MM.YYYY HH:MM'
                    else:
                        max_len = max((df_to_format[col_name].astype(str).map(len).max(), len(col_name)))
                        worksheet.column_dimensions[column_letter].width = max_len + 2

                if sheet_name == 'Cost Calculation':
                    # In this sheet, we apply a thick border only under 'TOTAL' rows
                    thick_bottom_border = Border(bottom=thick_side)
                    # Find the column index for 'Lineitem name' to reliably find TOTAL rows
                    try:
                        lineitem_name_col_idx = df_to_format.columns.get_loc('Lineitem name') + 1
                        for row_idx in range(2, worksheet.max_row + 1):
                            cell_value = worksheet.cell(row=row_idx, column=lineitem_name_col_idx).value
                            if cell_value == 'TOTAL':
                                for col_idx in range(1, worksheet.max_column + 1):
                                    worksheet.cell(row=row_idx, column=col_idx).border = thick_bottom_border
                    except KeyError:
                        logging.warning("Could not find 'Lineitem name' column to apply special formatting.")

                elif sheet_name in ['All Orders', 'Without Package Protection']:
                    # Original logic for sheets where we group by order name
                    thin_bottom_border = Border(bottom=thin_side)
                    thick_bottom_border = Border(bottom=thick_side)
                    for row_idx in range(2, worksheet.max_row + 1):
                        is_last = (row_idx == worksheet.max_row or
                                   worksheet.cell(row=row_idx, column=1).value != worksheet.cell(row=row_idx + 1, column=1).value)
                        border_to_apply = thick_bottom_border if is_last else thin_bottom_border
                        for col_idx in range(1, worksheet.max_column + 1):
                            worksheet.cell(row=row_idx, column=col_idx).border = border_to_apply

                elif sheet_name == 'Final Invoice':
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        for cell in row:
                            cell.border = thin_border

        logging.info(f"Successfully created Excel report: {output_filename}")
    except Exception as e:
        logging.error(f"Failed to create Excel report. Reason: {e}")
        sys.exit(1)

def prepare_report_sheets(report_df, cost_first_sku, cost_next_sku, cost_per_piece):
    """
    Prepares a dictionary of DataFrames for the multi-sheet Excel report.
    """
    protection_pattern = "Package protection|Shipping Protection"
    df_no_protection = report_df[~report_df['Lineitem name'].str.contains(protection_pattern, na=False)].copy()

    # Pass a copy to calculate_costs to avoid SettingWithCopyWarning
    df_with_costs = calculate_costs(report_df.copy(), cost_first_sku, cost_next_sku, cost_per_piece)

    # Define and select the final columns for the 'Cost Calculation' report
    cost_report_cols = [
        'Name', 'Fulfilled at', 'Lineitem quantity', 'Lineitem name', 'Lineitem sku',
        'Piece Cost', 'SKU Cost', 'Line Total Cost'
    ]
    # Ensure all required columns exist in the dataframe before selection
    for col in cost_report_cols:
        if col not in df_with_costs.columns:
            df_with_costs[col] = 0 # or some other default

    # Fill NaN values in the relevant columns before processing
    final_cost_cols = ['Piece Cost', 'SKU Cost', 'Line Total Cost']
    df_with_costs[final_cost_cols] = df_with_costs[final_cost_cols].fillna(0)

    df_costs_for_report = df_with_costs[cost_report_cols]

    # The transform function now handles its own sorting robustly.
    df_costs_transformed = transform_cost_df_for_reporting(df_costs_for_report)

    # The invoice summary is calculated on the original, untransformed cost data
    df_invoice = create_invoice_summary(df_with_costs, cost_first_sku, cost_next_sku, cost_per_piece)

    return {
        'All Orders': report_df,
        'Without Package Protection': df_no_protection,
        'Cost Calculation': df_costs_transformed,
        'Final Invoice': df_invoice
    }

def main():
    """
    Main function to orchestrate the order processing workflow.
    """
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    logging.info("Starting Shopify Order Processor...")

    # --- User Input ---
    start_date = get_date_from_user("Enter the start date (DD.MM.YYYY): ")
    end_date = get_date_from_user("Enter the end date (DD.MM.YYYY): ")
    print("\nPlease enter the cost tariffs:")
    cost_first_sku = get_tariff_from_user("Enter the cost for the first SKU: ")
    cost_next_sku = get_tariff_from_user("Enter the cost for each subsequent SKU: ")
    cost_per_piece = get_tariff_from_user("Enter the cost per piece: ")

    # --- Data Processing ---
    input_filename = "orders_export.csv"
    logging.info(f"Input file: {input_filename}")
    logging.info(f"Processing orders from {start_date.strftime('%d.%m.%Y')} to {end_date.strftime('%d.%m.%Y')}")

    source_df = load_and_validate_csv(input_filename)
    filtered_df = filter_by_date_range(source_df, start_date, end_date)

    final_columns = [
        'Name', 'Fulfilled at', 'Fulfillment Status', 'Financial Status',
        'Created at', 'Lineitem quantity', 'Lineitem name', 'Lineitem sku',
        'Lineitem fulfillment status'
    ]
    existing_columns = [col for col in final_columns if col in filtered_df.columns]
    report_df = filtered_df[existing_columns]

    # --- Report Generation ---
    if not report_df.empty:
        sheets_data = prepare_report_sheets(report_df, cost_first_sku, cost_next_sku, cost_per_piece)
        prompt_message = "Enter the desired name for the output Excel file (e.g., report.xlsx).\nPress Enter to use a default name: "
        output_filename_from_user = input(prompt_message)

        if output_filename_from_user:
            output_filename = output_filename_from_user
            if not output_filename.endswith('.xlsx'):
                output_filename += '.xlsx'
        else:
            current_date = datetime.now().strftime("%Y-%m-%d")
            output_filename = f"processed_orders_{current_date}.xlsx"
            logging.info(f"No filename provided. Using default: {output_filename}")
        create_excel_report(sheets_data, output_filename)
    else:
        logging.info("Script finished. No orders to process into an Excel file.")

if __name__ == "__main__":
    main()
